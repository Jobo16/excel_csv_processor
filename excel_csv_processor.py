#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel文件处理器 - 完整工作流程
功能：
1. 将Excel文件转换为CSV文件（处理合并单元格）
2. 合并CSV文件中的指定列到新的题目信息列
3. 使用相对路径，从input文件夹读取，输出到output文件夹

"""

import pandas as pd
import os
import csv
from pathlib import Path
from openpyxl import load_workbook

class ExcelCSVProcessor:
    """Excel和CSV文件处理器"""
    
    def __init__(self, base_dir=None):
        """
        初始化处理器
        
        Args:
            base_dir (str): 基础目录，默认为当前脚本所在目录
        """
        if base_dir is None:
            self.base_dir = Path(__file__).parent
        else:
            self.base_dir = Path(base_dir)
        
        self.input_dir = self.base_dir / "input"
        self.output_dir = self.base_dir / "output"
        
        # 确保输出目录存在
        self.output_dir.mkdir(exist_ok=True)
    
    def get_merged_cell_value(self, worksheet, row, col):
        """
        获取合并单元格的值
        
        Args:
            worksheet: openpyxl工作表对象
            row (int): 行号（1-based）
            col (int): 列号（1-based）
        
        Returns:
            str: 单元格的值
        """
        cell = worksheet.cell(row=row, column=col)
        
        # 检查是否为合并单元格
        for merged_range in worksheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # 获取合并单元格左上角的值
                top_left_cell = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                return top_left_cell.value
        
        return cell.value
    
    def fill_answer_instruction_column(self, csv_data):
        """
        处理"答题说明"列的空白单元格填充
        将有内容的单元格内容复制到下面相邻的空单元格中
        
        Args:
            csv_data (list): CSV数据列表
        
        Returns:
            list: 处理后的CSV数据列表
        """
        if not csv_data or len(csv_data) < 2:
            return csv_data
        
        header_row = csv_data[0]
        answer_instruction_col_index = -1
        for i, header in enumerate(header_row):
            if header and "答题说明" in str(header):
                answer_instruction_col_index = i
                break
        
        if answer_instruction_col_index == -1:
            return csv_data

        # 计算除"答题说明"列外，其他列的最大有效行数
        max_rows_other_cols = 0
        for r_idx, row in enumerate(csv_data):
            for c_idx, cell in enumerate(row):
                if c_idx != answer_instruction_col_index and cell and str(cell).strip():
                    if r_idx + 1 > max_rows_other_cols:
                        max_rows_other_cols = r_idx + 1

        # 如果其他列都为空，则不做任何操作
        if max_rows_other_cols == 0:
            max_rows_other_cols = len(csv_data)

        print(f"  找到答题说明列，索引: {answer_instruction_col_index}")
        print(f"  其他列最大行数: {max_rows_other_cols}")

        # 填充答题说明，但不超过其他列的最大行数
        last_instruction = ""
        for row_index in range(1, max_rows_other_cols):
            if row_index < len(csv_data):
                current_row = csv_data[row_index]
                
                # 确保当前行有足够的列
                while len(current_row) <= answer_instruction_col_index:
                    current_row.append("")
                
                current_cell_value = current_row[answer_instruction_col_index]
                
                if current_cell_value and str(current_cell_value).strip():
                    last_instruction = current_cell_value
                elif last_instruction:
                    current_row[answer_instruction_col_index] = last_instruction

        # 如果"答题说明"列的行数超过了其他列，进行截断
        return [row[:len(header_row)] for row in csv_data[:max_rows_other_cols]]
    
    def convert_xlsx_to_csv(self, xlsx_path, csv_path):
        """
        将Excel文件转换为CSV文件，处理合并单元格
        
        Args:
            xlsx_path (Path): Excel文件路径
            csv_path (Path): 输出CSV文件路径
        
        Returns:
            bool: 转换是否成功
        """
        try:
            print(f"  正在转换: {xlsx_path.name}")
            
            # 加载工作簿
            workbook = load_workbook(xlsx_path, data_only=True)
            
            # 处理第一个工作表
            worksheet = workbook.active
            
            # 获取工作表的最大行数和列数
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            # 创建数据列表
            csv_data = []
            
            # 遍历所有行和列
            for row in range(1, max_row + 1):
                row_data = []
                for col in range(1, max_col + 1):
                    # 获取单元格值（处理合并单元格）
                    cell_value = self.get_merged_cell_value(worksheet, row, col)
                    
                    # 处理None值
                    if cell_value is None:
                        cell_value = ""
                    else:
                        cell_value = str(cell_value)
                    
                    row_data.append(cell_value)
                
                csv_data.append(row_data)
            
            # 处理"答题说明"列的空白单元格填充
            csv_data = self.fill_answer_instruction_column(csv_data)
            
            # 写入CSV文件
            with open(csv_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerows(csv_data)
            
            print(f"  ✅ 转换成功: {xlsx_path.name} -> {csv_path.name}")
            return True
            
        except Exception as e:
            print(f"  ❌ 转换失败 {xlsx_path.name}: {str(e)}")
            return False
    
    def merge_columns_to_question_info(self, csv_path, output_path):
        """
        合并CSV文件中的指定列到新的题目信息列
        
        Args:
            csv_path (Path): 输入CSV文件路径
            output_path (Path): 输出CSV文件路径
        
        Returns:
            bool: 处理是否成功
        """
        try:
            print(f"  正在合并列: {csv_path.name}")
            
            # 读取CSV文件，使用UTF-8 BOM编码
            df = pd.read_csv(csv_path, encoding='utf-8-sig')
            
            # 检查必要的列是否存在
            required_columns = ['材料内容', '*题目类型', '*题干', '*正确答案']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                print(f"  ⚠️ 缺少必要的列: {missing_columns}，跳过列合并")
                return False
            
            # 定义选项列（A到K）
            option_columns = [f'选项{chr(65+i)}' for i in range(11)]  # 选项A到选项K
            
            # 创建新的题目信息列
            question_info_list = []
            
            for index, row in df.iterrows():
                # 收集非空的内容
                info_parts = []
                
                # 添加材料内容（如果非空且不是NaN）
                if pd.notna(row['材料内容']) and str(row['材料内容']).strip():
                    info_parts.append(f"材料内容：\n{str(row['材料内容']).strip()}")

                # 添加答题说明（如果存在且非空）
                if '答题说明' in df.columns and pd.notna(row['答题说明']) and str(row['答题说明']).strip():
                    info_parts.append(f"答题说明：\n{str(row['答题说明']).strip()}")
                
                # 添加题目类型
                if pd.notna(row['*题目类型']) and str(row['*题目类型']).strip():
                    info_parts.append(f"题目类型：\n{str(row['*题目类型']).strip()}")
                
                # 添加题干
                if pd.notna(row['*题干']) and str(row['*题干']).strip():
                    info_parts.append(f"题干：\n{str(row['*题干']).strip()}")
                
                # 添加正确答案
                if pd.notna(row['*正确答案']) and str(row['*正确答案']).strip():
                    info_parts.append(f"正确答案：\n{str(row['*正确答案']).strip()}")
                
                # 添加非空的选项
                options = []
                for option_col in option_columns:
                    if option_col in df.columns and pd.notna(row[option_col]) and str(row[option_col]).strip():
                        option_letter = option_col[-1]  # 获取选项字母（A, B, C等）
                        options.append(f"{option_letter}. {str(row[option_col]).strip()}")
                
                if options:
                    info_parts.append(f"选项：\n{chr(10).join(options)}")
                
                # 合并所有部分
                question_info = "\n\n".join(info_parts) if info_parts else ""
                question_info_list.append(question_info)
            
            # 添加新的题目信息列
            df['题目信息'] = question_info_list
            
            # 保存结果，使用UTF-8 BOM编码
            df.to_csv(output_path, index=False, encoding='utf-8-sig')
            
            print(f"  ✅ 列合并成功: {len(df)} 行数据 -> {output_path.name}")
            return True
            
        except Exception as e:
            print(f"  ❌ 列合并失败 {csv_path.name}: {str(e)}")
            return False
    
    def process_all_files(self):
        """
        处理所有文件的完整工作流程
        1. 将input文件夹中的Excel文件转换为CSV
        2. 对转换后的CSV文件进行列合并
        3. 输出最终结果到output文件夹
        
        Returns:
            dict: 处理结果统计
        """
        print("🚀 开始Excel文件处理工作流程...")
        print(f"📁 输入目录: {self.input_dir}")
        print(f"📁 输出目录: {self.output_dir}")
        print("-" * 60)
        
        # 检查输入目录是否存在
        if not self.input_dir.exists():
            print(f"❌ 输入目录不存在: {self.input_dir}")
            return {"success": 0, "total": 0, "errors": ["输入目录不存在"]}
        
        # 获取所有Excel文件
        xlsx_files = list(self.input_dir.glob('*.xlsx'))
        
        if not xlsx_files:
            print(f"❌ 在输入目录中未找到任何Excel文件")
            return {"success": 0, "total": 0, "errors": ["未找到Excel文件"]}
        
        print(f"🔍 找到 {len(xlsx_files)} 个Excel文件")
        
        success_count = 0
        total_count = len(xlsx_files)
        errors = []
        
        for xlsx_file in xlsx_files:
            print(f"\n--- 处理文件: {xlsx_file.name} ---")
            
            try:
                # 步骤1: 转换Excel为CSV（临时文件）
                temp_csv_path = self.output_dir / f"temp_{xlsx_file.stem}.csv"
                
                if not self.convert_xlsx_to_csv(xlsx_file, temp_csv_path):
                    errors.append(f"Excel转CSV失败: {xlsx_file.name}")
                    continue
                
                # 步骤2: 合并CSV列（最终输出文件）
                final_csv_path = self.output_dir / f"{xlsx_file.stem}_processed.csv"
                
                if self.merge_columns_to_question_info(temp_csv_path, final_csv_path):
                    success_count += 1
                    print(f"  ✅ 完整处理成功: {xlsx_file.name} -> {final_csv_path.name}")
                else:
                    # 如果列合并失败，保留转换后的CSV文件
                    temp_csv_path.rename(self.output_dir / f"{xlsx_file.stem}.csv")
                    success_count += 1
                    print(f"  ⚠️ 仅完成Excel转CSV: {xlsx_file.name} -> {xlsx_file.stem}.csv")
                
                # 清理临时文件
                if temp_csv_path.exists():
                    temp_csv_path.unlink()
                    
            except Exception as e:
                error_msg = f"处理文件 {xlsx_file.name} 时出错: {str(e)}"
                errors.append(error_msg)
                print(f"  ❌ {error_msg}")
        
        print("\n" + "=" * 60)
        print(f"🎉 处理完成! 成功处理 {success_count}/{total_count} 个文件")
        
        if errors:
            print(f"\n⚠️ 遇到 {len(errors)} 个错误:")
            for error in errors:
                print(f"  - {error}")
        
        return {
            "success": success_count,
            "total": total_count,
            "errors": errors
        }

def main():
    """
    主函数
    """
    # 创建处理器实例
    processor = ExcelCSVProcessor()
    
    # 执行完整的处理流程
    result = processor.process_all_files()
    
    # 显示最终结果
    if result["success"] == result["total"] and result["total"] > 0:
        print(f"\n🎊 所有文件处理成功！")
    elif result["success"] > 0:
        print(f"\n✅ 部分文件处理成功，请检查错误信息")
    else:
        print(f"\n❌ 没有文件处理成功，请检查输入文件和错误信息")

if __name__ == "__main__":
    main()